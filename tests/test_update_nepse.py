import sys
import unittest
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

from update_nepse import DataError, NoTradingData, SourceSnapshot, is_nepse_scheduled_day, merge_row, parse_market_page, parse_price_page, should_publish_snapshot
from import_history_xlsx import load_workbook_rows


MARKET_HTML = """
<div>As of 2026-08-24</div>
<table><thead><tr><th>Index</th><th>Open</th><th>High</th><th>Low</th><th>Close</th><th>Point Change</th><th>% Change</th><th>Turnover</th></tr></thead>
<tbody><tr><td>NEPSE Index</td><td>2,600</td><td>2,650</td><td>2,580</td><td>2,620</td><td>20</td><td>0.7</td><td>4,100,000,000</td></tr></tbody></table>
<table><thead><tr><th>Sub Index</th><th>Open</th><th>High</th><th>Low</th><th>Close</th><th>Point</th><th>% Change</th><th>Turnover</th></tr></thead>
<tbody><tr><td>Banking SubIndex</td><td>1,400</td><td>1,420</td><td>1,390</td><td>1,410</td><td>10</td><td>0.7</td><td>1,000,000,000</td></tr></tbody></table>
"""

PRICE_HTML = """
<table><thead><tr><th>S.No</th><th>Symbol</th><th>Conf.</th><th>Open</th><th>High</th><th>Low</th><th>Close</th><th>LTP</th><th>VWAP</th><th>Vol</th></tr></thead>
<tbody>
<tr><td>1</td><td>ADBL</td><td>-</td><td>300</td><td>305</td><td>295</td><td>302</td><td>302</td><td>301</td><td>12,345</td></tr>
<tr><td>2</td><td>ACBL</td><td>-</td><td>0</td><td>0</td><td>0</td><td>0</td><td>0</td><td>0</td><td>0</td></tr>
</tbody></table>
"""


class UpdateNepseTests(unittest.TestCase):
    def test_parse_market_page_preserves_index_and_subindex_fields(self):
        date, entities = parse_market_page(MARKET_HTML)
        self.assertEqual(date, "2026-08-24")
        self.assertEqual(entities["Index"]["Open"], 2600)
        self.assertEqual(entities["Index"]["Turnover"], 4100000000)
        self.assertEqual(entities["Banking SubIndex"]["Close"], 1410)

    def test_parse_price_page_keeps_valid_symbol_ohlc_and_volume(self):
        symbols = parse_price_page(PRICE_HTML)
        self.assertEqual(list(symbols), ["ADBL"])
        self.assertEqual(
            symbols["ADBL"],
            {"Open": 300, "High": 305, "Low": 295, "Close": 302, "Volume": 12345},
        )

    def test_merge_row_preserves_existing_fields_and_sorts_history(self):
        old = [
            {"Date": "2026-08-23", "Index - Close": 2600},
            {"Date": "2026-08-24", "Index - Close": 2610},
        ]
        updated = merge_row(old, {"Date": "2026-08-24", "Index - Close": 2620})
        self.assertEqual([row["Date"] for row in updated], ["2026-08-23", "2026-08-24"])
        self.assertEqual(updated[-1]["Index - Close"], 2620)

    def test_merge_row_does_not_erase_existing_fields(self):
        old = [{"Date": "2026-08-24", "Index - Close": 2610, "ADBL-Close": 302}]
        updated = merge_row(old, {"Date": "2026-08-24", "Index - Close": 2620})
        self.assertEqual(updated[0]["Index - Close"], 2620)
        self.assertEqual(updated[0]["ADBL-Close"], 302)

    def test_xlsx_importer_reads_wide_rows_and_normalizes_dates(self):
        from openpyxl import Workbook
        from tempfile import TemporaryDirectory

        with TemporaryDirectory() as directory:
            path = Path(directory) / "history.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Date", "Index - Open", "Index - High", "Index - Low", "Index - Close", "Index - Turnover", "ADBL-Open", "ADBL-High", "ADBL-Low", "ADBL-Close", "ADBL-Volume"])
            sheet.append(["2026-08-23", 2600, 2620, 2590, 2610, 1000000, 300, 305, 295, 302, 1234])
            sheet.append(["2026-08-24", 2610, 2630, 2600, 2620, 1100000, 302, 308, 300, 307, 2345])
            workbook.save(path)
            rows = load_workbook_rows(path)

        self.assertEqual([row["Date"] for row in rows], ["2026-08-23", "2026-08-24"])
        self.assertEqual(rows[-1]["Index - Close"], 2620)
        self.assertEqual(rows[-1]["ADBL-Volume"], 2345)

    def test_invalid_market_page_is_rejected(self):
        with self.assertRaises(NoTradingData):
            parse_market_page("<html><body>No market table</body></html>")

    def test_market_page_without_explicit_date_is_rejected(self):
        html = MARKET_HTML.replace("As of 2026-08-24", "Market summary")
        with self.assertRaises(NoTradingData):
            parse_market_page(html)

    def test_nepse_schedule_matches_apps_script_monday_to_friday(self):
        # Python weekday(): Monday=0 ... Friday=4, Saturday=5, Sunday=6.
        self.assertTrue(is_nepse_scheduled_day(date(2026, 8, 24)))
        self.assertTrue(is_nepse_scheduled_day(date(2026, 8, 28)))
        self.assertFalse(is_nepse_scheduled_day(date(2026, 8, 29)))
        self.assertFalse(is_nepse_scheduled_day(date(2026, 8, 30)))

    def test_stale_source_date_is_not_published(self):
        snapshot = SourceSnapshot(
            trading_date="2026-08-26",
            row={"Date": "2026-08-26", "Index - Close": 2600},
            symbol_count=300,
            index_count=18,
        )
        self.assertFalse(should_publish_snapshot(snapshot, date(2026, 8, 27)))


if __name__ == "__main__":
    unittest.main()
