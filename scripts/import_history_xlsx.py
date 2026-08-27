#!/usr/bin/env python3
"""Seed the repository wide-row dataset from a Google Sheets XLSX export.

The workbook is expected to have one sheet whose first row contains the same
wide headers consumed by the SMC Chart, for example:
Date, Index - Open, ..., ADBL-Open, ADBL-High, ...
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_ROOT = Path(__file__).resolve().parents[1]
DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
INDEX_METRICS = {"Open", "High", "Low", "Close", "Turnover"}
SYMBOL_METRICS = {"Open", "High", "Low", "Close", "Volume"}


def iso_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    if "T" in text:
        text = text.split("T", 1)[0]
    return text[:10]


def json_value(value: Any) -> Any:
    if value is None or value == "":
        return ""
    if isinstance(value, (datetime, date)):
        return iso_date(value)
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "—", "–", "N/A", "NA"}:
        return ""
    text = text.replace("%", "")
    try:
        number = float(text)
    except ValueError:
        return str(value).strip()
    return int(number) if number.is_integer() else number


def load_workbook_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if not workbook.sheetnames:
        raise ValueError("Workbook has no worksheets")
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    try:
        raw_headers = list(next(rows))
    except StopIteration as exc:
        raise ValueError("Worksheet is empty") from exc
    headers = [str(value).strip() if value is not None else "" for value in raw_headers]
    if not headers or headers[0].lower() != "date":
        raise ValueError("The first worksheet column must be Date")
    if len(set(h for h in headers if h)) != len([h for h in headers if h]):
        raise ValueError("Worksheet contains duplicate non-empty headers")

    by_date: dict[str, dict[str, Any]] = {}
    for row in rows:
        values = list(row)
        if len(values) < len(headers):
            values.extend([None] * (len(headers) - len(values)))
        trading_date = iso_date(values[0] if values else None)
        if not DATE_RE.fullmatch(trading_date):
            continue
        output: dict[str, Any] = {"Date": trading_date}
        for index, header in enumerate(headers[1:], start=1):
            if not header:
                continue
            value = json_value(values[index] if index < len(values) else None)
            if value != "":
                output[header] = value
        if len(output) > 1:
            # If a date is duplicated in the workbook, the later non-empty row
            # wins field-by-field rather than erasing earlier columns.
            existing = by_date.setdefault(trading_date, {"Date": trading_date})
            existing.update(output)
    return [by_date[key] for key in sorted(by_date)]


def load_json_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        return []
    return [row for row in data if isinstance(row, dict) and row.get("Date")]


def merge_rows(history: list[dict[str, Any]], current: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    for row in history + current:
        trading_date = iso_date(row.get("Date"))
        if not DATE_RE.fullmatch(trading_date):
            continue
        existing = merged.setdefault(trading_date, {"Date": trading_date})
        existing.update({key: value for key, value in row.items() if value not in (None, "")})
    return [merged[key] for key in sorted(merged)]


def columns(rows: list[dict[str, Any]]) -> list[str]:
    keys = {key for row in rows for key in row}
    index_keys = sorted(
        (key for key in keys if " - " in key),
        key=lambda key: (key.rsplit(" - ", 1)[0] != "Index", key),
    )
    symbol_keys = sorted(key for key in keys if " - " not in key and key != "Date")
    return ["Date", *index_keys, *symbol_keys]


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows({field: row.get(field, "") for field in fieldnames} for row in rows)


def symbol_count(row: dict[str, Any]) -> int:
    names = set()
    for key in row:
        match = re.match(r"^(.+)-(Open|High|Low|Close|Volume)$", key, re.I)
        if match:
            names.add(match.group(1))
    return len(names)


def index_count(row: dict[str, Any]) -> int:
    return len({key.rsplit(" - ", 1)[0] for key in row if " - " in key})


def write_outputs(root: Path, rows: list[dict[str, Any]], source_name: str) -> None:
    data_dir = root / "data"
    fieldnames = columns(rows)
    latest = rows[-1]
    write_json(data_dir / "nepse_ohlc.json", rows)
    write_csv(data_dir / "nepse_ohlc.csv", rows, fieldnames)
    write_json(data_dir / "latest.json", latest)
    write_csv(data_dir / "latest.csv", [latest], fieldnames)
    write_json(data_dir / "history" / f"{latest['Date']}.json", latest)

    index_fields = [field for field in fieldnames if field == "Date" or " - " in field]
    fast_rows = [{field: row[field] for field in index_fields if field in row} for row in rows]
    write_json(data_dir / "nepse_ohlc_fast.json", fast_rows)

    manifest = {
        "source": {"historical_seed": source_name},
        "updated_at_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "latest_trading_date": latest["Date"],
        "history_rows": len(rows),
        "latest_symbol_count": symbol_count(latest),
        "latest_index_count": index_count(latest),
        "format": "SMC Toolkit wide rows: Date, Index - Open, SYMBOL-Open, ...",
    }
    write_json(data_dir / "manifest.json", manifest)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", type=Path, help="Google Sheets XLSX export")
    parser.add_argument("--root", type=Path, default=DEFAULT_ROOT, help="Repository root")
    args = parser.parse_args()

    imported = load_workbook_rows(args.xlsx)
    if not imported:
        raise SystemExit("No dated historical rows found in the workbook")
    current = load_json_rows(args.root / "data" / "nepse_ohlc.json")
    # Use the existing live row as a base, then overlay the user's historical
    # workbook. This keeps the imported Sheet authoritative for overlapping
    # dates while retaining any live fields absent from the workbook.
    merged = merge_rows(current, imported)
    write_outputs(args.root, merged, args.xlsx.name)
    print(
        f"Imported {len(imported)} workbook dates; wrote {len(merged)} total dates "
        f"from {merged[0]['Date']} to {merged[-1]['Date']}"
    )
    print(f"Latest row: {symbol_count(merged[-1])} symbols, {index_count(merged[-1])} indices")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
